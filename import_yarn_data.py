"""
One-time migration: read 'SK  - Yarn Tracker.xlsx' and populate
yarn_colors + yarn_transactions tables.

Run from the project root:
    python import_yarn_data.py

Requires: openpyxl, sqlalchemy, app package accessible in PYTHONPATH.
The DATABASE_URL must be set (via .env or environment) before running.
"""

import sys
from datetime import date, datetime

import openpyxl

# ── ensure app package is importable ──────────────────────────────────────────
import os
sys.path.insert(0, os.path.dirname(__file__))

from app.database import SessionLocal, engine, Base  # noqa: E402
# Import all models so SQLAlchemy's class registry can resolve every relationship
from app.models import (  # noqa: F401
    activity, client, design, lead, project, project_files,
    quotation, social_post, system_log, task, user, yarn,
)
from app.models.yarn import YarnColor, YarnTransaction
from app.utils.time import now_ist

EXCEL_FILE = "SK  - Yarn Tracker.xlsx"


def _normalize_code(raw) -> str | None:
    """Convert a cell value to a clean color code string."""
    if raw is None:
        return None
    if isinstance(raw, float):
        return str(int(raw))
    return str(raw).strip()


def _parse_date(raw) -> date | None:
    """Parse a date string 'DD-MM-YYYY' or return None."""
    if not raw:
        return None
    if isinstance(raw, (date, datetime)):
        return raw.date() if isinstance(raw, datetime) else raw
    s = str(raw).strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    print(f"  WARNING: could not parse date {s!r}")
    return None


def main():
    print(f"Loading {EXCEL_FILE} …")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    # ── 1. Read Yarn Master ────────────────────────────────────────────────────
    ws_master = wb["Yarn Master"]
    master_rows = list(ws_master.iter_rows(values_only=True))

    colors_data = []
    for row in master_rows[2:]:  # skip 2 header rows
        if len(row) < 2 or row[1] is None:
            continue
        code = _normalize_code(row[1])
        if not code:
            continue
        opening = float(row[2]) if len(row) > 2 and row[2] is not None else 0.0
        colors_data.append((code, opening))

    print(f"  Yarn Master: {len(colors_data)} colors found")

    # ── 2. Read Yarn Transactions ──────────────────────────────────────────────
    ws_tx = wb["Yarn Transaction"]
    tx_rows = list(ws_tx.iter_rows(values_only=True))

    # Rows 3–906 in Excel = index 2..905 (0-indexed)
    tx_data = []
    for row in tx_rows[2:906]:
        sr_no = row[4] if len(row) > 4 else None
        if sr_no is None:
            continue
        color_raw = row[5] if len(row) > 5 else None
        date_raw  = row[6] if len(row) > 6 else None
        stock_in  = row[7] if len(row) > 7 else None
        stock_out = row[8] if len(row) > 8 else None

        code = _normalize_code(color_raw)
        if not code:
            continue

        tx_date = _parse_date(date_raw)
        if tx_date is None:
            tx_date = now_ist().date()

        if stock_in is not None and float(stock_in) > 0:
            tx_data.append((code, "in",  float(stock_in),  tx_date))
        elif stock_out is not None and float(stock_out) > 0:
            tx_data.append((code, "out", float(stock_out), tx_date))

    print(f"  Yarn Transactions: {len(tx_data)} transactions found")

    # ── 3. Insert into DB ──────────────────────────────────────────────────────
    db = SessionLocal()
    try:
        existing_colors = db.query(YarnColor.color_code).all()
        existing_codes = {r[0] for r in existing_colors}

        if existing_codes:
            print(f"\n  WARNING: {len(existing_codes)} colors already exist in DB.")
            confirm = input("  Overwrite / skip duplicates? (yes/no): ").strip().lower()
            if confirm != "yes":
                print("  Aborted.")
                return

        # Insert colors
        color_map: dict[str, int] = {}
        inserted_colors = 0
        skipped_colors = 0
        now = now_ist()

        for code, opening in colors_data:
            if code in existing_codes:
                # Update existing record's id into map
                existing = db.query(YarnColor).filter(YarnColor.color_code == code).first()
                if existing:
                    color_map[code] = existing.id
                skipped_colors += 1
                continue
            yc = YarnColor(color_code=code, opening_stock=opening, created_at=now)
            db.add(yc)
            db.flush()
            color_map[code] = yc.id
            existing_codes.add(code)
            inserted_colors += 1

        db.commit()
        print(f"\n  Colors: inserted={inserted_colors}, skipped={skipped_colors}")

        # Re-load any that were skipped (so map is complete)
        if skipped_colors:
            for yc in db.query(YarnColor).all():
                color_map[yc.color_code] = yc.id

        # Insert transactions
        existing_tx_count = db.query(YarnTransaction).count()
        if existing_tx_count > 0:
            print(f"\n  WARNING: {existing_tx_count} transactions already exist in DB.")
            confirm = input("  Delete all and re-import? (yes/no): ").strip().lower()
            if confirm == "yes":
                db.query(YarnTransaction).delete()
                db.commit()
                print("  Cleared existing transactions.")
            else:
                print("  Skipping transaction import.")
                return

        inserted_tx = 0
        skipped_tx = 0
        for code, tx_type, qty, tx_date in tx_data:
            color_id = color_map.get(code)
            if color_id is None:
                print(f"  WARNING: color code {code!r} not in DB, skipping transaction")
                skipped_tx += 1
                continue
            db.add(YarnTransaction(
                color_id=color_id,
                transaction_type=tx_type,
                quantity=qty,
                date=tx_date,
                project_id=None,
                notes=None,
                created_at=now,
            ))
            inserted_tx += 1

        db.commit()
        print(f"  Transactions: inserted={inserted_tx}, skipped={skipped_tx}")

        # ── 4. Summary ────────────────────────────────────────────────────────
        total_colors = db.query(YarnColor).count()
        total_tx = db.query(YarnTransaction).count()
        print(f"\n  Done. DB now has {total_colors} colors and {total_tx} transactions.")

    except Exception as e:
        db.rollback()
        print(f"\nERROR: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
